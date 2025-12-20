from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required
from .decorators import group_required, multi_group_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.contrib import messages 

# MODELOS
from app.models import Trabajador, Cargo, Carga_familiar, Area, Departamento, Sexo_trabajador

# DESCARGAS
import csv
import openpyxl

# HOME
@login_required
def home(req):
    return render(req, 'home.html')

# AUTENTICACIÓN
def signup(req):
    form = UserCreationForm(req.POST)  # Instanciando el form
    
    if req.method == 'POST':
        try:
            if req.POST['password1'] == req.POST['password2']:
                try:
                    # registrar Usuario
                    user = User.objects.create_user(username=req.POST['username'], password=req.POST['password1'])
                    user.save()

                    # iniciando sesion
                    login(req, user)

                    return redirect('dashboard') 
                except IntegrityError:
                    return render(req, 'signup.html', {
                        'form': form,
                        'error': 'El usuario ya existe',
                    })
            else:
                return render(req, 'signup.html', {
                        'form': form,
                        'error': 'Las contraseñas no coinciden.',
                    })
        except:
            return redirect('signup')
    
    elif req.method == 'GET':
        return render(req, 'signup.html', {'form': form})  # Pasamos el formulario al template

def signin(req):
    form = AuthenticationForm()

    if req.method == 'GET':
        return render(req, 'signin.html', {'form': form})
    else:
        try:
            user = authenticate(req, username=req.POST['username'], password=req.POST['password'])
            if user is None:
                return render(req, 'signin.html', {'form': form, 'error': 'Usuario no registrado'})
            else:

                login(req, user)
                return redirect('dashboard')
        except:
            return redirect('signin')

@login_required
def signout(req):
    logout(req)
    return render(req, 'home.html')
    # return redirect('home.html')

# TRANSVERSALES
@login_required
def dashboard(req):
    # Verificar si el usuario es superuser
    if req.user.is_superuser:
        es_jefe_rrhh = True
        es_personal_rrhh = True
        es_trabajador = True
    else:
        # Verificar si el usuario pertenece a ciertos grupos
        es_jefe_rrhh = req.user.groups.filter(name='Jefe RRHH').exists()
        es_personal_rrhh = req.user.groups.filter(name='Personal RRHH').exists()
        es_trabajador = req.user.groups.filter(name='Trabajador').exists()
    context = {
        'es_jefe_rrhh': es_jefe_rrhh,
        'es_personal_rrhh': es_personal_rrhh,
        'es_trabajador': es_trabajador,
    }
    return render(req, 'dashboard.html', context)
    # return render(req, 'dashboard.html', {'notificaciones_pendientes': True})

@login_required
@multi_group_required(['Jefe RRHH', 'Personal RRHH'])
def informe_trabajadores(req):
    # 1. Recuperamos toda la data necesaria para los filtros
    areas = Area.objects.all()
    departamentos = Departamento.objects.all()
    cargos = Cargo.objects.all()
    sexos = Sexo_trabajador.choices

    # 2. La enviamos al template en el contexto
    context = {
        'areas': areas,
        'departamentos': departamentos,
        'cargos': cargos,
        'sexos': sexos,
    }
    return render(req, 'informe-trabajadores.html', context)


@login_required
@multi_group_required(['Jefe RRHH', 'Personal RRHH'])
def datos_filtrados(req):
    
    # 1. Obtener parámetros de filtro (GET)
    sexo = req.GET.get('sexo', '').strip()
    cargo_id = req.GET.get('cargo', '').strip()
    depto_id = req.GET.get('departamento', '').strip()
    area_id = req.GET.get('area', '').strip()
    
    # Variables de control para exportación
    export_csv = req.GET.get('export_csv', '').strip()
    export_excel = req.GET.get('export_excel', '').strip()

    # 2. Aplicar Filtros a la QuerySet base
    trabajadores = Trabajador.objects.all()

    if sexo:
        trabajadores = trabajadores.filter(sexo_trabajador=sexo)
    if cargo_id:
        trabajadores = trabajadores.filter(id_cargo=cargo_id)
    if depto_id:
        trabajadores = trabajadores.filter(id_cargo__id_departamento=depto_id)
    if area_id:
        trabajadores = trabajadores.filter(id_cargo__id_departamento__id_area=area_id)

    # ==========================================
    # LÓGICA DE EXPORTACIÓN (CSV)
    # ==========================================
    if export_csv == '1':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="trabajadores-filtrados.csv"'
        
        # BOM para que Excel reconozca tildes y ñ en UTF-8
        response.write(u'\ufeff'.encode('utf8'))
        
        writer = csv.writer(response, delimiter=';')
        
        # Encabezado
        writer.writerow(['RUT', 'Nombre', 'Apellidos', 'Sexo', 'Cargo', 'Departamento', 'Área', 'Fecha Ingreso'])
        
        # Datos
        for t in trabajadores:
            writer.writerow([
                t.rut_trabajador,
                t.nombre_trabajador,
                t.apellidos_trabajador,
                t.get_sexo_trabajador_display(),
                t.id_cargo.nombre_cargo,
                t.id_cargo.id_departamento.nombre_departamento,
                t.id_cargo.id_departamento.id_area.nombre_area,
                t.fecha_ingreso_trabajador.strftime("%d-%m-%Y")
            ])
        return response

    # ==========================================
    # LÓGICA DE EXPORTACIÓN (EXCEL .xlsx)
    # ==========================================
    if export_excel == '1':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="trabajadores-filtrados.xlsx"'

        # Crear libro de trabajo y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trabajadores"

        # Encabezados
        headers = ['RUT', 'Nombre', 'Apellidos', 'Sexo', 'Cargo', 'Departamento', 'Área', 'Fecha Ingreso']
        ws.append(headers)

        # Datos
        for t in trabajadores:
            ws.append([
                t.rut_trabajador,
                t.nombre_trabajador,
                t.apellidos_trabajador,
                t.get_sexo_trabajador_display(),
                t.id_cargo.nombre_cargo,
                t.id_cargo.id_departamento.nombre_departamento,
                t.id_cargo.id_departamento.id_area.nombre_area,
                t.fecha_ingreso_trabajador.strftime("%d-%m-%Y")
            ])

        # Ajuste básico de ancho de columnas (Opcional, solo estética)
        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 20

        wb.save(response)
        return response

    # ==========================================
    # RENDERIZADO NORMAL (Si no descarga nada)
    # ==========================================
    filtros = {
        'sexo': sexo,
        'cargo': cargo_id,
        'departamento': depto_id,
        'area': area_id
    }

    return render(req, 'datos-filtrados.html', {
        'trabajadores': trabajadores,
        'filtros': filtros,
    })

@login_required
def informe_horas_trabajadas(req):
    return render(req, 'informe-horas-trabajadas.html')

# ========================
# PERFIL PERSONAL RRHH
# ========================
@login_required
@multi_group_required(['Personal RRHH'])
def llenar_ficha_trabajador(req):
    cargos = Cargo.objects.all()
    
    if req.method == 'POST':
        # 1. Recolección de datos
        campos_requeridos = [
            'rut_trabajador', 'nombre_trabajador', 'apellidos_trabajador',
            'direccion_trabajador', 'sexo_trabajador', 'id_cargo'
        ]
        
        campos_trabajador = {}
        error_msg = None # Usamos una variable local para controlar el error
        
        for campo in campos_requeridos:
            valor = req.POST.get(campo, '').strip()
            if not valor:
                error_msg = f'El campo "{campo}" es obligatorio.'
                break
            campos_trabajador[campo] = valor

        # Si hay error básico
        if error_msg:
            messages.error(req, error_msg) # Encolamos el mensaje de error
            return render(req, 'llenar-ficha-trabajador.html', {
                'cargos': cargos,
                'valores_anteriores': req.POST
            })
        
        # 2. Validación de Cargo
        try:
            id_cargo_str = campos_trabajador.pop('id_cargo')
            cargo_instance = Cargo.objects.get(id=int(id_cargo_str))
        except (Cargo.DoesNotExist, ValueError):
            messages.error(req, 'El cargo seleccionado no es válido.')
            return render(req, 'llenar-ficha-trabajador.html', {
                'cargos': cargos,
                'valores_anteriores': req.POST
            })
        
        # 3. Crear el Trabajador
        try:
            nuevo_trabajador = Trabajador.objects.create(
                id_cargo=cargo_instance,
                **campos_trabajador
            )
            
            # MENSAJE DE ÉXITO
            messages.success(req, f'¡Trabajador {nuevo_trabajador.nombre_trabajador} creado correctamente!')
            
            # Redirigir al informe (listado) para ver que se agregó
            return redirect('informe_trabajadores') 
            
        except Exception as e:
            # MENSAJE DE ERROR DE BD (ej. Rut duplicado si tuvieras unique=True)
            messages.error(req, f'Error al guardar en base de datos: {e}')
            return render(req, 'llenar-ficha-trabajador.html', {
                'cargos': cargos,
                'valores_anteriores': req.POST
            })

    else:
        return render(req, 'llenar-ficha-trabajador.html', {'cargos': cargos})
    
# PERFIL TRABAJADOR
@login_required
def llenar_ficha_carga_familiar(req):
    if req.method == 'POST':
        # Obtener datos del formulario
        nombre = req.POST.get('nombre_carga_familiar', '').strip()
        parentesco = req.POST.get('parentesco_carga_familiar', '').strip()
        rut = req.POST.get('rut_carga_familiar', '').strip()
        sexo = req.POST.get('sexo_carga_familiar', '').strip()

        # # Validación mínima (puedes mejorarla si necesitas)
        # if not (nombre and parentesco and rut and sexo):
        #     messages.error(req, "Por favor, complete todos los campos obligatorios.")
        #     return render(req, 'llenar-ficha-carga-familiar.html')

        try:
            # Encontrar al trabajador por nombre y apellido del usuario
            user = req.user
            trabajador = Trabajador.objects.get(nombre_trabajador=user.first_name, apellidos_trabajador=user.last_name)

            cargas_query = Carga_familiar.objects.filter(id_trabajador=trabajador)
            cargas_familiares = [
                {'nombre': carga.nombre_carga_familiar, 'relacion': carga.parentesco_carga_familiar}
                for carga in cargas_query
            ]

            # Crear y guardar la carga familiar
            Carga_familiar.objects.create(
                id_trabajador=trabajador,
                nombre_carga_familiar=nombre,
                parentesco_carga_familiar=parentesco,
                rut_carga_familiar=rut,
                sexo_carga_familiar=sexo
            )

            messages.success(req, "Carga familiar registrada exitosamente.")
            return render(req, 'seleccionar-cargas.html', {'cargas_familiares': cargas_familiares})
            

        except Trabajador.DoesNotExist:
            messages.error(req, "No se encontró un trabajador asociado a tu usuario. Verifica tu nombre y apellido en el perfil.")
            return render(req, 'seleccionar-cargas.html', {'cargas_familiares': cargas_familiares})

    # Si es GET, solo renderiza el formulario
    return render(req, 'llenar-ficha-carga-familiar.html')

@login_required
def seleccionar_cargas_familiares(req):

    if req.method == 'GET':
        # Obtener el nombre y apellido del usuario autenticado
        user = req.user
        nombre = user.first_name
        apellido = user.last_name

        try:
            # Buscar el trabajador por nombre y apellido
            trabajador = Trabajador.objects.get(nombre_trabajador=nombre, apellidos_trabajador=apellido)
            # Obtener sus cargas familiares
            cargas_query = Carga_familiar.objects.filter(id_trabajador=trabajador)
            # Formatear como lista de diccionarios
            cargas_familiares = [
                {'nombre': carga.nombre_carga_familiar, 'relacion': carga.parentesco_carga_familiar}
                for carga in cargas_query
            ]
        except Trabajador.DoesNotExist:
            # Si no existe el trabajador, devolver lista vacía o mensaje
            cargas_familiares = []
    else:
        # Si no es POST, también podrías aplicar la misma lógica o dejar vacío
        user = req.user
        nombre = user.first_name
        apellido = user.last_name
        try:
            trabajador = Trabajador.objects.get(nombre_trabajador=nombre, apellidos_trabajador=apellido)
            cargas_query = Carga_familiar.objects.filter(id_trabajador=trabajador)
            cargas_familiares = [
                {'nombre': carga.nombre_carga_familiar, 'relacion': carga.parentesco_carga_familiar}
                for carga in cargas_query
            ]
        except Trabajador.DoesNotExist:
            cargas_familiares = []

    return render(req, 'seleccionar-cargas.html', {'cargas_familiares': cargas_familiares})

@login_required
def seleccionar_contactos_emergencia(req):
    contactos_emergencia = [
    {'nombre': 'María López', 'relacion': 'Esposa'},
    {'nombre': 'Carlos Ruiz', 'relacion': 'Hijo'},
    {'nombre': 'Ana Torres', 'relacion': 'Madre'},
    {'nombre': 'Luis Pérez', 'relacion': 'Hermano'},
    {'nombre': 'Alberta Jara', 'relacion': 'Hija'}
    ]
    return render(req, 'seleccionar-contactos.html', {'contactos_emergencia': contactos_emergencia})

def modificar_datos_personales(req):
    return render(req, 'modificar-datos-personales.html')

def marcado(req):
    return render(req, 'marcado.html')